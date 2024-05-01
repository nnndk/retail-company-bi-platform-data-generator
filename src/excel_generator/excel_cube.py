from src.excel_generator.excel_dimension import ExcelDimension
from src.excel_generator.excel_fact_dimension import ExcelFactDimension
from src.excel_generator.excel_dimension_item import ExcelDimensionItem
from src.tools.common_tools import *

from datetime import datetime, timedelta
import openpyxl
import openpyxl.utils as openpyxl_utils
from openpyxl.styles import Font, Alignment


class ExcelCube:
    """
    This class is a representation of an OLAP-cube.
    An instance of this class consists of dimensions (1 or more), time series, and facts (also 1 or more, one fact = [])
    """
    _base_path = '../generated_files/'

    def __init__(self, cube_name: str, dimensions: list[ExcelDimension], facts: list[ExcelFactDimension]):
        """
        Constructor
        :param cube_name: The name of OLAP-cube. Also, it is a default Excel filename
        :param dimensions: List of cube dimensions (ExcelDimension)
        :param facts: List of cube facts (ExcelFactDimension)
        """
        self.cube_name = cube_name
        self.dimensions = dimensions
        self.facts = facts

    @staticmethod
    def create_excel_dim(dim_name: str, fields: list[str], dim_items_list: dict[str, list[str]]) -> ExcelDimension:
        """
        Static method that provides an easier way to create ExcelDimension
        :param dim_name: The name of new ExcelDimension
        :param fields: Fields of this ExcelDimension
        :param dim_items_list: Values of each dimension field
        :return: new ExcelDimension instance
        """
        dim_items = []

        for i in range(len(dim_items_list[fields[0]])):
            dim_item = {}

            for field in fields:
                dim_item[field] = dim_items_list[field][i]

            dim_items.append(ExcelDimensionItem(i, dim_item))

        return ExcelDimension(dim_name, fields, dim_items)

    def generate(self, start_date: datetime, end_date: datetime, min_rows: int, max_rows: int, filename='') -> None:
        """
        This method generates an Excel file with given dimensions, facts, and time series (start_date and end_date).
        It is created some number of rows (random from min_rows to max_rows) for each timestamp in the time series.
        Values of each dimension are randomly selected.
        Values of each fact are randomly generated accordingly to the params of the fact-object.
        Time series of this OLAP-cube is always daily.
        If a dimension has multiple fields, it will be named in output file as 'dim_name'.'field_name'
        :param start_date: The start of the time series
        :param end_date: The end of the time series
        :param min_rows: Min rows for each timestamp
        :param max_rows: Max rows for each timestamp
        :param filename: The name of the generated Excel-file. If filename = '' (default), then the filename will be
        the same as this ExcelCube (self.cube_name)
        :return: None
        """
        if filename == '':
            filename = self.cube_name

        start = start_date.date()
        end = end_date.date()
        current_date = start

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        row, col = 1, 1
        sheet.cell(row=1, column=1).value = 'Дата'
        sheet.cell(row=1, column=1).font = Font(bold=True)
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        for i in range(len(self.dimensions)):
            dim = self.dimensions[i]

            for field in dim.fields:
                col += 1

                if field == 'value':
                    sheet.cell(row=1, column=col).value = f'{dim.name}'
                else:
                    sheet.cell(row=1, column=col).value = f'{dim.name}.{field}'

                sheet.cell(row=1, column=col).font = Font(bold=True)
                sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')

        for fact in self.facts:
            col += 1
            sheet.cell(row=1, column=col).value = f'{fact.name}'
            sheet.cell(row=1, column=col).font = Font(bold=True)
            sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')

        while current_date <= end:
            cnt_rows = get_random_int(min_rows, max_rows)

            for i in range(0, cnt_rows):
                row += 1
                sheet.cell(row=row, column=1).value = current_date
                col = 1

                for dim in self.dimensions:
                    dim_item_id = get_random_int(0, len(dim.values) - 1)

                    for key in dim.values[dim_item_id].values:
                        col += 1
                        sheet.cell(row=row, column=col).value = dim.values[dim_item_id].values[key]
                        sheet.cell(row=row, column=col).alignment = Alignment(horizontal='left')

                for fact in self.facts:
                    col += 1
                    sheet.cell(row=row, column=col).value = fact.generate_fact_value()
                    sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')

            current_date += timedelta(days=1)

        ExcelCube._set_auto_width(sheet, col)
        workbook.save(ExcelCube._base_path + filename + '.xlsx')

    @staticmethod
    def _set_auto_width(sheet, last_col: int) -> None:
        """
        Fit width of columns to the text of their cells
        :param sheet: Excel sheet (workbook.active). It will be changed in this method
        :param last_col: The last col that should be fitted
        :return: None
        """
        for col in range(1, last_col + 1):
            max_length = 0

            for row in sheet.iter_rows(min_row=1, min_col=col, max_col=col, values_only=True):
                for cell in row:
                    if cell:
                        max_length = max(max_length, len(str(cell)))

            # Set the column width based on the maximum length of the content
            sheet.column_dimensions[openpyxl_utils.get_column_letter(col)].width = max_length + 2
